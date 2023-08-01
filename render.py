#!/usr/bin/python3.11
import argparse
import enum
import subprocess
import sys
import os
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName


class EnumAction(argparse.Action):
    # https://stackoverflow.com/a/60750535
    """
    Argparse action for handling Enums
    """

    def __init__(self, **kwargs):
        # Pop off the type value
        enum_type = kwargs.pop("type", None)

        # Ensure an Enum subclass is provided
        if enum_type is None:
            raise ValueError(
                "type must be assigned an Enum when using EnumAction")
        if not issubclass(enum_type, enum.Enum):
            raise TypeError("type must be an Enum when using EnumAction")

        # Generate choices from the Enum
        kwargs.setdefault("choices", tuple(e.value for e in enum_type))

        super(EnumAction, self).__init__(**kwargs)

        self._enum = enum_type

    def __call__(self, parser, namespace, values, option_string=None):
        # Convert value back into an Enum
        value = self._enum(values)
        setattr(namespace, self.dest, value)


class InputLabelSpacing(enum.Enum):
    LEFT = "left"
    RIGHT = "right"
    ABOVE = "above"
    BELOW = "below"
    NONE = "none"


parser = argparse.ArgumentParser()

parser.add_argument("src_file", type=str,
                    help="The XLSL source code file to compile")
parser.add_argument("output_cells", type=str,
                    help="The cell addresses in which to put the output of the XLSL main method, separated by commas or as a range")
parser.add_argument("-i", "--input-cells", type=str,
                    help="The cell addresses of the inputs to the XLSL main method, separated by commas or as a range")
parser.add_argument("-l", "--input-labels", type=InputLabelSpacing, default=InputLabelSpacing.NONE,
                    action=EnumAction, help="The positioning of input labels next to the input cells")
parser.add_argument("-o", "--output_file", type=str,
                    help="The name of the output Excel spreadsheet")
parser.add_argument("-t", "--template_file", type=str,
                    help="The name of a template Excel spreadsheet. If one is not provided, a blank workbook is created")

args = parser.parse_args()


def col_row(address):
    xy = coordinate_from_string(address)
    return (column_index_from_string(xy[0]), xy[1])


def addr(col, row):
    return get_column_letter(col)+str(row)


def parse_cell_lists(arg):
    cell_refs = []
    parts = arg.split(",")
    for part in parts:
        if ":" in part:
            start, end = part.split(":")
            start_x, start_y = col_row(start)
            end_x, end_y = col_row(end)
            for x in range(start_x, end_x+1):
                for y in range(start_y, end_y+1):
                    cell_refs.append(addr(x, y))
        else:
            cell_refs.append(part)
    return cell_refs


src_file = args.src_file
input_cells = parse_cell_lists(args.input_cells)
output_cells = parse_cell_lists(args.output_cells)

bin_path = "./build/compile"

cmdlineArgs = [bin_path, src_file, "y"]
for i in input_cells:
    cmdlineArgs.append(f"'Sheet'!{i}")

print(" ".join(cmdlineArgs))
proc = subprocess.run(cmdlineArgs,
                      stdout=subprocess.PIPE, stderr=subprocess.PIPE)

if (len(proc.stderr) > 0):
    print("Compiler failed with the following error message:\n\t" +
          proc.stderr.decode("utf-8").strip(), file=sys.stderr)
    exit(1)

response = list(x.strip()
                for x in proc.stdout.decode("utf-8").strip().split("\n"))

num_inputs = int(response[0])
input_labels = []
for i in range(num_inputs):
    input_labels.append(response[i+1])

formula = "="+response[num_inputs+1]
print(formula)
num_functions = int(response[num_inputs+2])
functions = []
for i in range(num_functions):
    functions.append((response[num_inputs+3+i*2],response[num_inputs+4+i*2]))

for function_name, function_formula in functions:
    print(f"{function_name}: {function_formula}", end="\n\n")


output_file = args.output_file
if output_file is None:
    output_file = os.path.basename(os.path.splitext(src_file)[0]) + ".xlsx"

wb = Workbook() if args.template_file is None else load_workbook(args.template_file)
ws = wb.active

for cell in output_cells:
    ws[cell] = formula

for cell, label in zip(input_cells, input_labels):
    label_x, label_y = col_row(cell)
    if args.input_labels == InputLabelSpacing.LEFT:
        label_x -= 1
    elif args.input_labels == InputLabelSpacing.RIGHT:
        label_x += 1
    elif args.input_labels == InputLabelSpacing.ABOVE:
        label_y -= 1
    elif args.input_labels == InputLabelSpacing.BELOW:
        label_y += 1
    else:
        continue
    ws[addr(label_x, label_y)] = label

for function_name, function_formula in functions[::-1]:
    defn = DefinedName(function_name, attr_text=function_formula)
    wb.defined_names[function_name] = defn

wb.save(output_file)
